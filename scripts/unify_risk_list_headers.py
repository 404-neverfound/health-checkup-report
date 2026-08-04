#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
风险清单表统一后处理：删除资产表首行空行 + 统一各表表头样式。

背景：
- 资产表导出文件第 1 行是空行、第 2 行才是表头，需删掉第 1 行让整表上移一行。
- 各风险清单表表头样式不统一（资产/事件/弱口令/漏洞/策略检查各不相同），
  以暴露面清单的表头样式为统一标准。

暴露面清单表头标准样式（实测 暴露面清单.xlsx 全部 sheet 一致）：
- 字体：微软雅黑 10pt、加粗、黑色 FF000000
- 背景：FFBDCFE7（浅蓝）、solid
- 对齐：水平居中、垂直居中、自动换行
- 边框：四边 thin

用法：
    python unify_risk_list_headers.py <table_type> <input.xlsx> [output.xlsx]

table_type: asset | incident | weakpwd | vuln | exposure | policy
- asset：删首空行（仅当第 1 行全空且第 2 行非空），再统一表头
- exposure：原样返回（它是标准样式来源，无需处理）
- 其它：仅统一表头样式，不改数据

只改表头行，不动数据行样式。
"""
import json
import os
import shutil
import sys

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from _path_helper import decode_argv
decode_argv()

# 暴露面清单表头标准样式
HEADER_FONT = Font(name='微软雅黑', size=10, bold=True, color='FF000000')
HEADER_FILL = PatternFill(fill_type='solid', fgColor='FFBDCFE7', bgColor='FFBDCFE7')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)
HEADER_SIDE = Side(style='thin')
HEADER_BORDER = Border(
    left=HEADER_SIDE,
    right=HEADER_SIDE,
    top=HEADER_SIDE,
    bottom=HEADER_SIDE,
)
# 与暴露面清单表头行行高保持一致
HEADER_ROW_HEIGHT = 18.0

VALID_TYPES = ('asset', 'incident', 'weakpwd', 'vuln', 'exposure', 'policy')


def normalize(value):
    return '' if value is None else str(value).strip()


def row_has_any_value(row):
    return any(normalize(cell.value) for cell in row)


def find_header_row(ws):
    """返回第一个非空行（作为表头行），找不到返回 1。"""
    for row in ws.iter_rows(min_row=1, max_row=10):
        if row_has_any_value(row):
            return row[0].row
    return 1


def remove_first_empty_row(ws):
    """资产表专用：第 1 行全空且第 2 行非空时，删除第 1 行让整表上移一行。"""
    first_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
    second_row = list(ws.iter_rows(min_row=2, max_row=2))[0]
    if not row_has_any_value(first_row) and row_has_any_value(second_row):
        ws.delete_rows(1, 1)
        return True
    return False


def unify_header_style(ws, header_row_idx):
    """把第 header_row_idx 行的所有有值单元格统一为暴露面表头样式，并冻结表头行。"""
    max_col = ws.max_column or 1
    for col in range(1, max_col + 1):
        cell = ws.cell(row=header_row_idx, column=col)
        if cell.value is None:
            # 空表头单元格也套同一样式，保证整行观感一致
            cell.value = None
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER

    ws.row_dimensions[header_row_idx].height = HEADER_ROW_HEIGHT
    # 冻结表头行：数据从表头下一行开始滚动
    ws.freeze_panes = f'A{header_row_idx + 1}'


def process_workbook(table_type, wb):
    """处理工作簿中的所有 sheet（事件表等主 sheet 在第一个；暴露面多 sheet 各表头统一）。"""
    for ws in wb.worksheets:
        if table_type == 'asset':
            remove_first_empty_row(ws)
        header_row_idx = find_header_row(ws)
        unify_header_style(ws, header_row_idx)


def main():
    if len(sys.argv) < 3:
        raise SystemExit(
            'Usage: unify_risk_list_headers.py <asset|incident|weakpwd|vuln|exposure|policy> <input.xlsx> [output.xlsx]'
        )

    table_type = sys.argv[1].lower()
    input_path = sys.argv[2]
    output_path = sys.argv[3] if len(sys.argv) >= 4 else input_path

    if table_type not in VALID_TYPES:
        raise SystemExit(f'不支持的表类型: {table_type}（仅支持 {" / ".join(VALID_TYPES)}）')
    if not os.path.isfile(input_path):
        raise SystemExit(f'输入文件不存在: {input_path}')

    # 暴露面是标准样式来源，不处理
    if table_type == 'exposure':
        if os.path.abspath(input_path) != os.path.abspath(output_path):
            shutil.copy2(input_path, output_path)
        print(json.dumps({'filePath': os.path.abspath(output_path)}, ensure_ascii=False))
        return

    wb = load_workbook(input_path)
    try:
        process_workbook(table_type, wb)
        # 写临时文件再替换，避免中途失败损坏原文件
        if os.path.abspath(input_path) != os.path.abspath(output_path):
            os.makedirs(os.path.dirname(os.path.abspath(output_path)) or '.', exist_ok=True)
            wb.save(output_path)
        else:
            tmp_path = output_path + '.tmp'
            wb.save(tmp_path)
            wb.close()
            os.replace(tmp_path, output_path)
    finally:
        try:
            wb.close()
        except Exception:
            pass

    print(json.dumps({'filePath': os.path.abspath(output_path)}, ensure_ascii=False))


if __name__ == '__main__':
    main()
