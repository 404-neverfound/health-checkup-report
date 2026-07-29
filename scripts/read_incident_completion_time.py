#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Read the "完成时间" column for a given incident from the incident Excel file."""

import json
import sys
from openpyxl import load_workbook

from _path_helper import decode_argv
decode_argv()


def normalize(value):
    """Normalize a cell value to string for header matching."""
    if value is None:
        return ""
    return str(value).strip().replace("\n", "").replace("\r", "").replace(" ", "")


def find_column(headers, candidates):
    """Find the first column index whose normalized header matches any candidate."""
    for i, h in enumerate(headers):
        nh = normalize(h)
        for c in candidates:
            if normalize(c) == nh:
                return i
    return None


def read_timestamp(value):
    """Try to parse a cell value as a datetime and return ISO string, or raw string."""
    if value is None:
        return None
    try:
        # openpyxl returns datetime objects for date cells
        from datetime import datetime
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        pass
    text = str(value).strip()
    return text if text else None


def main():
    if len(sys.argv) < 3:
        print(json.dumps({"completionTime": None, "error": "Usage: script.py <excel_path> <incident_id>"}))
        sys.exit(1)

    excel_path = sys.argv[1]
    incident_id = str(sys.argv[2]).strip()

    try:
        wb = load_workbook(excel_path, data_only=True)
        ws = wb.active

        header_row = next(ws.iter_rows(values_only=True))
        headers = [normalize(h) for h in header_row]

        # Find incident ID column (try multiple possible names)
        id_col = find_column(headers, [
            "事件ID", "事件编号", "incident_id", "uuId", "uuid",
            "安全事件编号", "服务事件编号", "serviceEventId"
        ])

        # Find completion time column
        completed_col = find_column(headers, ["完成时间"])

        if id_col is None:
            wb.close()
            print(json.dumps({"completionTime": None, "error": "Cannot find incident ID column"}))
            sys.exit(0)

        if completed_col is None:
            wb.close()
            print(json.dumps({"completionTime": None, "error": "Cannot find 完成时间 column"}))
            sys.exit(0)

        completion_time = None
        for row in ws.iter_rows(min_row=2, values_only=True):
            if id_col >= len(row):
                continue
            row_id = normalize(row[id_col])
            if row_id == incident_id:
                raw = row[completed_col] if completed_col < len(row) else None
                completion_time = read_timestamp(raw)
                break

        wb.close()
        print(json.dumps({"completionTime": completion_time}))
    except Exception as e:
        print(json.dumps({"completionTime": None, "error": str(e)}))


if __name__ == "__main__":
    main()
